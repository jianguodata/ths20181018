# coding=utf-8
import openpyxl
import datetime
import tushare as ts
import schedule
import pandas as pd
from time import sleep
from os import path

import get_stockfund
from zt_stock import ZtStock
from cta.Double_stop_clamp import DoubleStopClamp
import ma_34
import desi_1
import w33
import updatepositiondata

Message_PriceTime = datetime.time(8, 0, 0)  # 讯价时间（8：0：0-9：15：0）
Bidding_Time = datetime.time(9, 22, 45)  # 竞价时间（9：15-9：30）
Opening_Time = datetime.time(9, 29, 52)  # 开盘时间（9：30-11：30，13：00-15：00）
Closing_Time_5 = datetime.time(14, 55, 0)  # 收盘最后五分钟（14：55：0-15：0：0）
Closing_Time = datetime.time(15, 0, 0)  # 收盘时间（15：0：0--）
Data_Upload_Time = datetime.time(16, 30, 0)  # 数据上传时间（16：30：00--）
Data_End_Time = datetime.time(16, 55, 0)
Futures_Time = datetime.time(21, 0, 0)  # 期货交易时间（21：00-01：00）
Futures_End_Time = datetime.time(23, 30, 0)

today = datetime.datetime.now().strftime('%Y%m%d')

paths = '.\\daily_stock_data\\data_day'
path_pos = '.\\daily_stock_data\\pos'


def time_in_range(start, end, x):
    # 时间控制开关
    if start <= end:
        return start <= x <= end
    else:
        return start <= x or x <= end


def nowtime():
    print('\n')
    print(datetime.datetime.now().strftime('%F %T'))


def runp():
    # 打印持仓股票信息 Print position stock
    w33.pripos()
    # 顶部箭
    top_arrow = analyse_pos_stock()
    # 获取可用金额
    available_amount = w33.get_Availableamount()
    # 可用股票数量
    pos_stock = w33.get_PositionStock()
    # 持仓股票数量
    Pos_AllStock = w33.get_AllPositionStock()
    # 止损
    stop_loss = w33.StopLoss()
    # 止盈
    target_profit = w33.get_profit()
    # 初始资金
    initial_amount = 200000
    # 资金总额
    total_assets = w33.get_TotalAssets()

    # 资金使用率
    Fund_utilization_rate = available_amount / total_assets

    print(u'初始资金：', initial_amount, u'元')
    print('')
    print(u'总资产：', total_assets, u' 总收益：',
          round((total_assets - initial_amount) / initial_amount * 100, 2), '%')
    print('')
    print('---------策略运行第 {0} 天--------'.format(Strategy_operation()))
    print('')
    print(u'止损：', stop_loss)
    print('')
    print(u'止盈：', target_profit)
    print('')

    # 资金前50强策略
    buy_list = desi_1.get_buystock()

    # 集团主力资金
    # 止损
    for code_loss in stop_loss:
        w33.sell_Stock(code_loss)
    # 止盈卖出股票
    sslist = list(set(top_arrow).union(set(stop_loss)))

    tosell = [sell for sell in sslist if sell in pos_stock]
    for code in tosell:
        if not ma_34.ma_34_signal(code):
            w33.sell_Stock(code)
    print('顶部箭：', top_arrow)
    print('to_sell:', tosell)
    if len(open_low_data()) > 0:
        print('低开6%股票:',open_low_data())
        for low_stock in open_low_data():
            w33.sell_Stock(low_stock)
    # 取消撤单
    # w33.getData_Frozen()

    """
    buy_list:策略选出的股票,初始默认为None
    Pos_AllStock:持仓股票,初始默认为None
    """
    # 双停夹道战法买入
    if Fund_utilization_rate > 0.5:
        buy_double_code()

    if buy_list is None:
        tobuy = None
    else:
        tobuy = [code for code in buy_list if code not in Pos_AllStock]

    cFixedCash = 10000  # 固定现金额度，若为10000，意味着买入10000金额的股票
    if tobuy is not None:
        stock_num = int(available_amount // cFixedCash)
        if stock_num >= len(tobuy):
            for code in tobuy:
                price = ts.get_realtime_quotes(code)  # Single stock symbol
                df1 = price[['a5_p', 'price', 'pre_close']].astype(float)
                price_a5 = df1['a5_p'].values[0]
                stock_number = cFixedCash / price_a5 - cFixedCash / price_a5 % 100
                if stock_number >= 100:
                    w33.buy_stock(int(stock_number), code)
            print('to_buy:', tobuy)
        elif 1 <= stock_num < len(tobuy):
            for code in tobuy[0:stock_num]:
                price = ts.get_realtime_quotes(code)  # Single stock symbol
                df1 = price[['a5_p', 'price', 'pre_close']].astype(float)
                price_a5 = df1['a5_p'].values[0]
                stock_number = cFixedCash / price_a5 - cFixedCash / price_a5 % 100
                if stock_number >= 100:
                    w33.buy_stock(int(stock_number), code)
            print('to_buy:', tobuy[0:stock_num])


def One_Key_Account_Clearance():
    """
    当上证指数，创业板指数的趋势为跌势(熊市 )，账户开一键清仓
    """
    signal = desi_1.index_trend()
    if signal == 'bear_market':
        pos = w33.getData()
        Df_pos = pd.DataFrame(pos, columns=[u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'成本价', u'市价', u'盈亏比例'])
        Df_pos.set_index([u'证券代码'], inplace=True)
        position_stock = []
        for stock in pos:
            if stock[u'可用余额'] > 0:
                position_stock.append(stock[u'证券代码'])
        for stock_2 in position_stock:
            w33.sell_Stock(stock_2)
    elif signal == 'structural_bear_market':
        pass


def Strategy_operation():
    """
    策略运行第几天
    """
    Initial_date = '20200730'
    todaynow = datetime.datetime.now().strftime('%Y%m%d')
    date1 = datetime.datetime.strptime(todaynow[0:10], "%Y%m%d")
    date2 = datetime.datetime.strptime(Initial_date[0:10], "%Y%m%d")
    num = (date1 - date2).days
    # print(print('策略运行第 {0} 天'.format(num)))
    return num


def buy_double_code():
    """购买双停夹道股票"""
    Double_code = desi_1.double_stop_clamp_codes()
    buy_double = [code for code in Double_code if code not in w33.get_AllPositionStock()]
    for code in buy_double:
        price = ts.get_realtime_quotes(code)  # Single stock symbol
        df1 = price[['a5_p', 'price', 'pre_close']].astype(float)
        price_a5 = df1['a5_p'].values[0]
        perc = (df1['price'] - df1['pre_close']) / df1['pre_close']
        # perc = perc.values * 100
        cFixedCash = 10000  # 固定现金额度，若为10000，意味着买入10000金额的股票
        if w33.get_Availableamount() > cFixedCash:
            stock_number = cFixedCash / price_a5 - cFixedCash / price_a5 % 100
            if stock_number >= 100:
                w33.buy_stock(int(stock_number), code)


def holding_amount():
    """单次股票金额分析，本金超过1W第二天考虑卖出一半"""
    pos = w33.getData()
    df_pos = pd.DataFrame(pos, columns=[u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'成本价', u'市价', u'盈亏比例'])
    df_pos.set_index([u'证券代码'], inplace=True)
    excess_pos = df_pos[df_pos[u'可用余额'] * df_pos[u'成本价'] > 12000]
    if len(excess_pos) > 0:
        for index, stock in excess_pos.iterrows():
            # index 为证券代码
            cost_price = stock[u'成本价']
            cFixedCash = 10000  # 固定现金额度，若为10000，意味着买入10000金额的股票
            stock_number = cFixedCash / cost_price - cFixedCash / cost_price % 100
            sell_number = int(stock['可用余额']) - stock_number
            if stock[u'市价'] > stock[u'成本价']:
                w33.Sell_partStock(sell_number, index)


def analyse_pos_stock():
    """
    分析持仓股票,当天冲高回落超过一定比例卖出
    :return: 待卖股票
    """
    pos = w33.getData()
    Df_pos = pd.DataFrame(pos, columns=[u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'成本价', u'市价', u'盈亏比例'])
    Df_pos.set_index([u'证券代码'], inplace=True)
    positionstock = []
    sell_stock = []
    Open_low = []
    for stock in pos:
        if stock[u'可用余额'] > 0:
            positionstock.append(stock[u'证券代码'])
    if Bidding_Time <= date_now < Opening_Time:
        for code in positionstock:
            pd.read_csv('{0}\\{1}.csv'.format(path_pos, code[0:6]))

    if Opening_Time <= date_now <= Closing_Time_5:
        for stock in positionstock:
            postushare = ts.get_realtime_quotes(stock)  # 从tushare下载持仓数据对比
            posdata = postushare[['code', 'open', 'bid', 'price', 'high', 'low', 'pre_close']].astype(float)
            posdata_high = (posdata['high'] - posdata['price']) / posdata['price']
            posdata_high = posdata_high * 100
            # 开盘跌幅超过6%
            Decline = (posdata['open'] - posdata['pre_close']) / posdata['pre_close']
            Decline = Decline * 100
            if posdata_high.values > 5.5 and posdata['pre_close'].values > posdata['price'].values:
                sell_stock.append(stock)
            if posdata['open'].values == posdata['high'].values and posdata_high.values > 3.9:
                sell_stock.append(stock)
            if Decline.values < -6:
                Open_low.append(stock)
    return sell_stock


def open_low_data():
    """
    分析持仓股票,当天低开6个点直接卖出
    :return: 待卖股票
    """
    pos = w33.getData()
    Df_pos = pd.DataFrame(pos, columns=[u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'成本价', u'市价', u'盈亏比例'])
    Df_pos.set_index([u'证券代码'], inplace=True)
    positionstock = []
    Open_low = []
    for stock in pos:
        if stock[u'可用余额'] > 0:
            positionstock.append(stock[u'证券代码'])
    for stock in positionstock:
        postushare = ts.get_realtime_quotes(stock)  # 从tushare下载持仓数据对比
        posdata = postushare[['code', 'open', 'bid', 'price', 'high', 'low', 'pre_close']].astype(float)
        # 开盘跌幅超过6%
        Decline = (posdata['open'] - posdata['pre_close']) / posdata['pre_close']
        Decline = Decline * 100
        if Decline.values <= -4.5 and posdata['open'] > posdata['low']:
            Open_low.append(stock)
    return Open_low


def update_position_data():
    """持仓数据更新"""
    _path_pos = '.\\daily_stock_data\\pos\\Hold_pos.csv'
    _pos = w33.getData()
    pos_df = pd.DataFrame(_pos, columns=[
        u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'冻结数量', u'成本价', u'市价', u'盈亏比例', 'high'])
    # df.set_index([u'证券代码'], inplace=True)
    pos_df['明天卖出'] = pos_df['盈亏比例'] < 0.0
    if not path.exists(_path_pos):
        PositionStock = list(pos_df[u'证券代码'].values)
        pos_data = ts.get_realtime_quotes(PositionStock)
        pos_df['high'] = list(pos_data['high'].values)
        pos_df.to_csv(_path_pos, encoding="gbk")
    PositionStock = list(pos_df[u'证券代码'].values)     # 当天最新持仓股票代码
    data = pd.read_csv(_path_pos, encoding='gbk')
    code_list = data[u'证券代码']
    code_list = code_list.apply(lambda x: str(x).zfill(6))
    data[u'证券代码'] = code_list
    code_list = list(code_list)                     # 上一个交易日持仓股票代码
    sell_code = [s_code for s_code in code_list if s_code not in PositionStock]   # 当天卖出股票代码
    buy_code = [b_code for b_code in PositionStock if b_code not in code_list]    # 当天买入股票代码
    # print('sell_code', sell_code)
    # print('buy_code', buy_code)
    new_data = data[~data[u'证券代码'].isin(sell_code)]

    if len(buy_code):
        new_df = pos_df[pos_df[u'证券代码'].isin(buy_code)]
        buy_data = ts.get_realtime_quotes(buy_code)
        new_df['high'] = list(buy_data['high'].values)
        new_df = pd.DataFrame(new_df)
        new_data = new_data.append(new_df, ignore_index=False)

    new_data = pd.DataFrame(new_data, columns=[
        u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'冻结数量', u'成本价', u'市价', u'盈亏比例', 'high','明天卖出'])
    pos_data = ts.get_realtime_quotes(list(new_data[u'证券代码'].values))

    new_data['today_high'] = list(pos_data['high'].values)
    for index, row in new_data.iterrows():
        high = row['high']
        today_high = row['today_high']
        if float(high) < float(today_high):
            row['high'] = today_high
    del new_data['today_high']
    # print(new_data)
    new_data.to_csv(_path_pos, index=None, encoding='gbk')
    print('持仓数据更新完毕！！')


def flow_stock_pool():
    """更新股票池"""
    data = pd.read_csv(r'{}\Hold_pos.csv'.format(path_pos), encoding='GBK')
    ts_code = [str(x).zfill(6) for x in data[u'证券代码']]
    data[u'证券代码'] = [str(x).zfill(6) for x in data[u'证券代码']]
    # 剔除资金前100里持仓股票代码
    with open('stocklist.txt','r') as pl_1:
        arr_1 = pl_1.readlines()
        flow_50 = [x[0:6] for x in arr_1 if x[0:6] not in ts_code]
    with open('stocklist.txt', 'w') as fileobj:
        # 将名字列表，code拼接为一个字符串，并以换行符‘\n‘分隔
        fileobj.write('\n'.join(flow_50))
    # 剔除涨停板里持仓股票代码
    with open('zt_stocklist.txt','r') as pl_2:
        arr_2 = pl_2.readlines()
        zt = [x[0:6] for x in arr_2 if x[0:6] not in ts_code]
    with open('zt_stocklist.txt', 'w') as fileobj_2:
        fileobj_2.write('\n'.join(zt))


def net_values_stock():
    """每天自动更新资产净值"""
    # pro = ts.pro_api()
    net_file = r'.\daily_stock_data\net_values.xlsx'
    today_now = datetime.datetime.now().strftime("%Y%m%d")
    trade_date = pro.query('trade_cal', start_date=today_now, end_date=today_now)
    if not trade_date.is_open.values:
        return
    import os
    if not os.path.exists(net_file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A2'] = '序号'
        ws['B2'] = '日期'
        ws['C2'] = '市值'
        ws['D2'] = '总资产'
        ws['E2'] = '收益率'
        wb.save(net_file)
    workbook = openpyxl.load_workbook(net_file)
    sheet_names = workbook.sheetnames
    sheet0 = workbook[sheet_names[0]]
    sheet0_max_row = sheet0.max_row
    last_date = sheet0.cell(row=sheet0_max_row, column=2).value
    if last_date == today:
        return
    num = sheet0_max_row + 1
    sheet0["A{}".format(num)] = sheet0_max_row - 1
    sheet0["B{}".format(num)] = today
    sheet0["C{}".format(num)] = w33.stock_market_value()
    sheet0["D{}".format(num)] = w33.get_TotalAssets()
    sheet0['E{}'.format(num)] = (w33.get_TotalAssets() - 200000) / 200000 * 100
    workbook.save(net_file)
    print('产品净值更新完毕！！')


def stock_basic_pool():
    """更新上市公司股票列表"""
    data = pro.query('stock_basic', exchange='', list_status='L',
                     fields='ts_code,symbol,name,area,industry,market,exchange,list_date,is_hs')
    data.to_excel('./data/stock/stock_basic_pool.xlsx', 'w', encoding='GBK')


def Scheduled_Update_data():
    """定时更新"""
    week_day = datetime.datetime.now().strftime('%w')
    if 0 < int(week_day) < 6:
        print('开始更新策略！')
        DoubleStopClamp().Double_limit_up()                # 双停夹道
        get_stockfund.get_Top_50_inflows_data()            # 资金流入前50强
        updatepositiondata.update_position_data()          # 收盘更新持仓数据
        ZtStock().get_up_Data()                             # 涨停板
        net_values_stock()                                 # 资产净值表
        flow_stock_pool()                                  # 更新股票池
        stock_basic_pool()                                 # 更新上市公司股票列表


if __name__ == "__main__":
    pro = ts.pro_api()
    Start_update = True      # 定时更新
    schedule.every().day.at("18:00").do(Scheduled_Update_data)
    # 定时启动更新
    if Start_update:
        current = datetime.datetime.now()
        date_now = datetime.time(current.hour, current.minute, current.second)
        if date_now > datetime.time(18, 0, 0):
            Scheduled_Update_data()
            Start_update = False
    while True:
        weekday = datetime.datetime.now().strftime('%w')
        if 0 < int(weekday) < 6:
            try:
                current = datetime.datetime.now()
                nowtime() 
                schedule.run_pending()
                date_now = datetime.time(current.hour, current.minute, current.second)
                if Message_PriceTime < date_now <= Bidding_Time:
                    print('早盘消息收集准备中')
                    sleep(60)
                elif Bidding_Time < date_now <= Opening_Time:
                    qc = []
                    pos = w33.getData()
                    df = pd.DataFrame(pos, columns=[
                        u'证券代码', u'证券名称', u'股票余额', u'可用余额', u'冻结数量', u'成本价', u'市价', u'盈亏比例'])
                    df.set_index([u'证券代码'], inplace=True)
                    var = df[df[u'盈亏比例'] < 0]
                    var = var[var['可用余额'] > 0]
                    if not var.empty:
                        qc = var.index.values
                    if len(qc) > 0:
                        for code in qc:
                            if not ma_34.ma_34_signal(code):       # 先判断34天均线趋势是否向上,趋势为涨势暂不卖出
                                w33.sell_Stock(code)
                elif time_in_range(Opening_Time, Closing_Time, date_now):
                    """
                    # w33.init('niubility_20')
                    # w33.buy_stock(amou,code)
                    """
                    runp()
                    holding_amount()   # 持仓超过规定的数量，卖出多余的股票，保证帐户安全
                elif Futures_Time <= date_now <= Futures_End_Time:
                    print('期货交易时间')
                    sleep(120)
                    # runp()
                else:
                    print('休市时间')
                    # w33.kill('')
                    sleep(60)
                    # runp()
                    # holding_amount()

            except Exception as e:
                print(e)
                sleep(30)
        else:
            sleep(30)